import pdfplumber
import openpyxl as xl
import os

for arquivo in os.listdir("pdf"):
    
    if arquivo.lower().endswith(".pdf"):
        try:
            # abrindo o arquivo excel
            excel = xl.load_workbook("Base_de_dados_inspecoes.xlsx")
            aba = excel.active
            linha_inicial = len(aba['A']) + 1

            # ler o arquivo pdf e extrair os dados
            pdf = pdfplumber.open(f"pdf\\{arquivo}")
            pagina = pdf.pages[0]
            dados = pagina.extract_table()

            for indice, dado in enumerate(dados[1:], start=linha_inicial):
                if dado[0] == '':
                    pass

                else:
                    aba.cell(row=indice, column=1).value = dado[0]
                    aba.cell(row=indice, column=2).value = dado[1]
                    aba.cell(row=indice, column=3).value = dado[2]
                    aba.cell(row=indice, column=4).value = dado[3]
                    aba.cell(row=indice, column=5).value = dado[4]

            pdf.close()
            excel.save("Base_de_dados_inspecoes.xlsx")
            excel.close()

        except Exception as e:
            with open("log_erros.txt", "a") as log:
                log.write(f"Aconteceu um arro ao extrair informações do {arquivo}!\n")
                log.write("Erro: {e}")


    else:
        with open("log_erros.txt", "a") as log:
            log.write(f"O arquivo {arquivo} não é um pdf válido!\n")
